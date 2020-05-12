import asyncio, requests, re, os
from lxml import etree
from openpyxl import Workbook

# 协程获取美剧列表，并保存到excel

class TTMJ:

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '美剧'
        self.url = 'http://www.ttkmj.org/c/qhkh'
        self.num = 1

    async def qhkx(self, url='http://www.ttkmj.org/c/qhkh'):
        r = requests.get(url).text
        html = etree.HTML(r)
        title = html.xpath('//li[@class="subject-item"]/div[@class="info"]/h2/a')
        for i in title:
            name = i.xpath('./text()')[0]
            title_name = re.sub('[/·？！：.:?!~～\r\t\n 下载]', '', name)
            title_url = i.xpath('./@href')[0]
            print(title_name, title_url)
            self.ws.cell(row=self.num, column=1).value = title_name
            self.ws.cell(row=self.num, column=2).value = title_url
            self.num += 1
        # print('end')

        extend = html.xpath('//div[@class="paginator"]/ul[@class="pagelist"]/div[@class="page_navi"]/a[@href]/text()')[-2]
        next = html.xpath('//div[@class="paginator"]/ul[@class="pagelist"]/div[@class="page_navi"]/a/@href')[-2]

        if extend == ' 下一页 ':
            await self.qhkx(next)
        else:
            self.wb.save(os.path.join(os.path.expanduser("~"), 'Desktop', '美剧列表.xlsx'))
            pass

    async def run(self):
        loop = asyncio.get_event_loop()
        tasks = [loop.create_task(self.qhkx(self.url))]
        for task in tasks:
            await task

if __name__ == "__main__":
    ttmj = TTMJ()
    loop = asyncio.get_event_loop()
    loop.run_until_complete(ttmj.run())
